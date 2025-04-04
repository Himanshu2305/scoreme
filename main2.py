import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
import re
from openpyxl.utils.exceptions import IllegalCharacterError

def clean_text(text):
    """Removes illegal characters that cannot be written to Excel."""
    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)  # Removes non-printable ASCII characters

def save_tables_to_excel(tables, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    for table in tables:
        for row in table:
            try:
                cleaned_row = [clean_text(str(cell)) for cell in row]  # Clean each cell
                ws.append(cleaned_row)
            except IllegalCharacterError:
                print("Skipped a row due to illegal characters.")
        ws.append([])  # Add a blank line between tables

    wb.save(output_path)
    print(f"Data saved successfully to {output_path}")


def clean_text(text):
    """Removes illegal characters that cannot be written to Excel."""
    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)  # Removes non-printable ASCII characters

def save_tables_to_excel(tables, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    for table in tables:
        for row in table:
            try:
                cleaned_row = [clean_text(str(cell)) for cell in row]  # Clean each cell
                ws.append(cleaned_row)
            except IllegalCharacterError:
                print("Skipped a row due to illegal characters.")
        ws.append([])  # Add a blank line between tables

    wb.save(output_path)
    print(f"Data saved successfully to {output_path}")



def clean_text(text):
    """Removes illegal characters that cannot be written to Excel."""
    return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)  # Removes non-printable ASCII characters


def save_tables_to_excel(tables, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    for table in tables:
        for row in table:
            try:
                cleaned_row = [clean_text(str(cell)) for cell in row]  # Clean each cell
                ws.append(cleaned_row)
            except IllegalCharacterError:
                print("Skipped a row due to illegal characters.")
        ws.append([])  # Add a blank line between tables

    wb.save(output_path)
    print(f"Data saved successfully to {output_path}")


def extract_text_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    all_text = []

    for page in doc:
        text = page.get_text("text")
        all_text.append(text)

    return all_text


def detect_tables(text_pages):
    tables = []

    for page_text in text_pages:
        lines = page_text.split("\n")
        table = []

        for line in lines:
            if re.search(r'\s{2,}', line):
                table.append([col.strip() for col in re.split(r'\s{2,}', line)])

        if table:
            tables.append(table)

    return tables


def save_tables_to_excel(tables, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Tables"

    for table in tables:
        for row in table:
            ws.append(row)
        ws.append([])

    wb.save(output_path)


def process_pdf(pdf_path, output_excel):
    text_pages = extract_text_from_pdf(pdf_path)
    tables = detect_tables(text_pages)
    save_tables_to_excel(tables, output_excel)



pdf_file = "sample.pdf"  # here you enter your file
output_excel = "extracted_tables.xlsx"
process_pdf(pdf_file, output_excel)
